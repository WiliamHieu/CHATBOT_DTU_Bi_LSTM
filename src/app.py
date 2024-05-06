import time
import streamlit as st
import tensorflow as tf
from chatbot_response import classify
from datetime import datetime
from openpyxl import Workbook, load_workbook


# Function to log the conversation
def log_conversation(question, answer):
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    with open("conversation_log.txt", "a", encoding="utf-8") as file:
        file.write(f"{timestamp} - Question: {question}\n")
        file.write(f"{timestamp} - Answer: {answer}\n")

    # Append to Excel file
    if not st.session_state.get('excel_initialized', False):
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Question', 'Answer'])
        st.session_state['excel_initialized'] = True
    else:
        wb = load_workbook('conversation_log.xlsx')
        ws = wb.active

    ws.append([timestamp, question, answer])
    wb.save('conversation_log.xlsx')


with st.sidebar.container():
    st.title("Chatbot hỗ trợ sinh viên")
    st.markdown("""
    <a href="https://is.duytan.edu.vn/">
        <img src="https://duytan.edu.vn/images/icon/logo_dtu_footer.png" alt="Duy Tan University Logo" width="200">
    </a>
    """, unsafe_allow_html=True)
    st.write("")
    st.write("Địa chỉ : Phòng 204-205, Tòa nhà G, 120 Hoàng Minh Thảo, Đà Nẵng")
    st.write("Điện thoại : (+84) (236) 710 99 22")
    st.write("FB: facebook.com/khoadaotaoquocte")
    st.write("")
    st.write("")
    st.image("qr.jpg")
    st.write("")
    st.write("")
    st.write("Tác giả: Võ Đình Hiếu (26311118231)")
    st.write("Điện thoại:(+84) 94 189 1998")



# Chat interface
st.caption("Học viên vui lòng nhập câu hỏi tại đây: ")
input = st.empty()
click_clear = st.button('CLEAR', key=3)
human = input.text_input('', key=1)
st.write("Chào em, chương trình tư vấn tự động (chatbot) của trường Đại học Duy Tân có thể hỗ trợ gì cho em?")

if human:
    st.write(f"Học viên: {human}")
    tag, result = classify(human)
    st.write(f"Bot: {result}")
    # Log the conversation
    log_conversation(human, result)

    if tag == 'phan_nan':
        st.write(".....")
        time.sleep(3)
        st.write("Tư vấn viên: Chào em, Thầy/Cô có thể giúp gì cho em")

if click_clear:
    human = input.text_input('', value='', key=2)
